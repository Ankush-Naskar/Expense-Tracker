import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


# === MODULE FOR INPUT ===
def Input(categories):

    i = 1
    category_list = []
    for element in categories:
        category_list.append(f"{i}. {element}")
        i += 1

    #  === INPUT CATEGORY ===
    try:
        category_index = int(input(f"Enter category NO. - {category_list} : "))
        if category_index < 1 or category_index > len(categories):
            print("\nCategory number out of range.")
            return None
    except ValueError:
        print("\nPlease enter a valid category number.")
        return None
    
    # === INPUT PRODUCT NAME ===
    product_name = input("Product Name : ")

    # === INPUT AMOUNT ===
    try:
        amount = float(input("Enter the amount : "))
    except ValueError:
        print("\nPlease enter a valid numeric amount. ")
        return None
    
    # === Date ====
    entry_date = datetime.now().strftime("%d-%m-%Y")

    # === category ====
    category = categories[category_index - 1]

    # === DATA LIST === 
    expense = {
        "Date": [entry_date],
        "Category": [category],
        "Product":[product_name],
        "Amount": [amount]
        }

    return expense

# === MODULE FORE CREATE DATA SHEET ===
def save_as_xl(expense):
    df = pd.DataFrame(expense)

    excel_file = r"D:\my projects\price_tracker\expenses.xlsx"
    try:
        book = load_workbook(excel_file)
        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
    except PermissionError:
        print("Permission denied: Close the Excel file before running this script.")
    except FileNotFoundError:
        df.to_excel(excel_file, index=False)

